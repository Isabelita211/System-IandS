import typing
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QStackedWidget, QMessageBox, QTableWidgetItem
from PyQt5.uic import loadUi
from PyQt5.QtCore import QDate, QDateTime,QTime, QTimer
from mysql.connector import connect, Error
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
import io
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import sys

# Obtener la fecha actual
fecha_actual = QDateTime.currentDateTime().date()
hora_actual = QTime.currentTime()

def extraer_informacion():
    url = 'https://www.bcv.org.ve'
    respuesta = requests.get(url, verify=True)
    soup = BeautifulSoup(respuesta.text, 'html.parser')
    informacion = soup.find_all('div', {'class': 'col-sm-6 col-xs-6 centrado'})[4]
    return float(informacion.text.strip().replace(',', '.'))

class InicioSesion(QMainWindow): #Carga la interfaz de inicio de sesion
    def __init__(self):
        super(InicioSesion, self).__init__()
        loadUi('Login.ui',self)
        self.setFixedHeight(700) #Ajustar el tamaño de la ventana
        self.setFixedWidth(1366)

class MenuPrin(QMainWindow): #Carga la interfaz de Menu Principal
    def __init__(self):
        super(MenuPrin, self).__init__()
        loadUi('MenuPrin.ui',self)

class Invent(QMainWindow): #Carga la interfaz de Inventario
    def __init__(self):
        super(Invent, self).__init__()
        loadUi('Inventario.ui', self)

class Producto(QMainWindow): #Carga la interfaz de Productos
    def __init__(self):
        super(Producto, self).__init__()
        loadUi('Productos.ui',self)

class Nomina(QMainWindow): #Carga la interfaz de Nomina
    def __init__(self):
        super(Nomina, self).__init__()
        loadUi('Nomina.ui',self)

class Proveedores(QMainWindow): #Cargar la interfaz de Proveedores
    def __init__(self):
        super(Proveedores, self).__init__()
        loadUi('Proveedores.ui',self)

class Venta(QMainWindow): #Cargar la interfaz de Venta
    def __init__(self):
        super(Venta, self).__init__()
        loadUi('Venta.ui',self)

class Empleados(QMainWindow): #Cargar la interfaz de Empleados
    def __init__(self):
        super(Empleados, self).__init__()
        loadUi('Empleados.ui',self)

class Datos(QMainWindow): #Cargar la interfaz de Datos Nomina
    def __init__(self):
        super(Datos,self).__init__()
        loadUi('DatosNomina.ui',self)

class Pagar(QMainWindow): #Cargar la interfaz de Forma de Pago
    def __init__(self):
        super(Pagar,self).__init__()
        loadUi('Pagar.ui',self)

class Auditoria(QMainWindow): #Cargar la interfaz de Auditoria
    def __init__(self):
        super(Auditoria, self).__init__()
        loadUi('Auditoria.ui', self)

class App(QMainWindow):
    def __init__(self):
        super(App, self).__init__()
        self.widget = QStackedWidget(self) #Creación de apilación de widgets
        self.Inicio = InicioSesion()
        self.Menu = MenuPrin()
        self.Invent = Invent()
        self.Productos = Producto()
        self.Nomina = Nomina()
        self.Prov = Proveedores()
        self.Venta = Venta()
        self.Emple = Empleados()
        self.Datos = Datos()
        self.Pago = Pagar()
        self.Audi = Auditoria()

        self.widget.addWidget(self.Inicio)
        self.widget.addWidget(self.Menu)
        self.widget.addWidget(self.Invent)
        self.widget.addWidget(self.Productos)
        self.widget.addWidget(self.Nomina)
        self.widget.addWidget(self.Prov)
        self.widget.addWidget(self.Venta)
        self.widget.addWidget(self.Emple)
        self.widget.addWidget(self.Datos)
        self.widget.addWidget(self.Pago)
        self.widget.addWidget(self.Audi)

        self.setCentralWidget(self.widget) #Asignar Widget central

        self.widget.setCurrentWidget(self.Inicio) #Asignar Widget actual

        self.conexiones()

        self.is_updating = False

        self.timer = QTimer()
        self.timer.singleShot(1000, self.actualizar_tablas1) #Llama a actualizar_tablas cada 3 seg
        self.timer.singleShot(60000,self.actualizar_hora) #Llama actualizar hora cada 1 min

    def conectarBD(self):
        try:
            connection = connect( #Conexion con la ubicacion de la base de datos
                host="localhost",
                user="root",
                password="",
                database="system_iands"
            )
            return connection
        except Error as e:
            print(f"Error al conectar a la base de datos: {e}") #Error si no encuentra la base de datos
            return None

    def login(self):
        clave_usuario = self.Inicio.Password.text().strip() #Agarra la informacion del QLine
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        if len(clave_usuario):
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor() #Busqueda de la informacion en la tabla
                query = "SELECT COUNT(*) FROM usuarios WHERE Clave = %s"
                cursor.execute(query, (clave_usuario,))
                result = cursor.fetchone()[0]
                if result > 0: #Si la clave es correcta
                    mensaje.setIcon(QMessageBox.Information)
                    mensaje.setText('Iniciando sesion')
                    mensaje.exec_()
                    self.widget.setCurrentWidget(self.Menu)
                else: #Si la clave es incorrecta
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('Contraseña incorrecta')
                    mensaje.exec_()
                connection.close()
            else: #Error si no encuentra la base de datos
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()
        else: #Campo del QLineText vacía
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Ingrese su contraseña')
            mensaje.exec_()
        self.Inicio.Password.clear() #Limpia el cuadro de texto

    def buscar_product(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor() #Busqueda de la informacion en la tabla
            query = "SELECT NomProd FROM inventario WHERE NomProd = %s"
            try:
                cursor.execute(query, (self.Productos.BuscarProd.text(),))
                result = cursor.fetchone()
                if result is not None:
                    self.Productos.NombreProd.setText(result[0])
                    self.Productos.NombreProd.setEnabled(False) #Deshabilitar cuadro de texto
                    mensaje.setIcon(QMessageBox.Information)
                    mensaje.setText('Producto encontrado')
                    mensaje.exec_()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('Producto no encontrado')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error al buscar el Producto: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        self.Productos.BuscarProd.clear() #Limpia el cuadro de texto

    def buscar_rif(self):
        rif_empresa = self.Productos.Rif.text().strip()
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        if len(rif_empresa):
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor() #Busqueda de la informacion en la tabla
                query = "SELECT NomEmpr FROM proveedores WHERE Rif = %s"
                try:
                    cursor.execute(query, (rif_empresa,))
                    result = cursor.fetchone()
                    if result is not None: # Verificar si result es None
                        self.Productos.Proveedor.setText(result[0])
                        self.Productos.Rif.setEnabled(False)
                        mensaje.setIcon(QMessageBox.Information)
                        mensaje.setText('Rif encontrado')
                        mensaje.exec_()
                    else:
                        mensaje.setIcon(QMessageBox.Warning)
                        mensaje.setText('Rif no encontrado')
                        mensaje.exec_()
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error al buscar el Rif: {e}')
                    mensaje.exec_()
                connection.close()
            else:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Ingrese un Rif válido')
            mensaje.exec_()

    def calculate_iva(self):
        # Agarrar los datos de los cuadros de texto
        PrecioTT = self.Productos.PrecioTT.text().strip()
        Cantidad = self.Productos.Num1.text().strip()

        # Agarrar los datos de los ComboBox
        IVA = float(self.Productos.IVA.currentText().strip('%'))

        # Agarrar los datos de los RadioButton y calcular el IVA
        IVAR = 'Si' if self.Productos.IVAsi.isChecked() else 'No'

        if IVAR == 'Si':
            PrecioIva = (float(PrecioTT) / float(Cantidad)) * (IVA / 100)
            PrecioExt = (float(PrecioTT) / float(Cantidad))
            PrecioUU = PrecioIva + PrecioExt
            PrecioUni = self.Productos.PrecioUnid.setText("{:.2f}".format(PrecioUU))
            PrecioIVa = self.Productos.PrecioIVA.setText("{:.2f}".format(PrecioIva))
        else:
            PrecioUU = (float(PrecioTT) / float(Cantidad))
            PrecioUni = self.Productos.PrecioUnid.setText("{:.2f}".format(PrecioUU))
            self.Productos.PrecioIVA.clear()

    def add_info(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        #Agarrar los datos de los cuadros de texto
        RifEmpr = self.Productos.Rif.text().strip()
        Prov = self.Productos.Proveedor.text().strip()
        Code = self.Productos.CodeProduct.text().strip()
        NumFact = self.Productos.Factura.text().strip()
        NumControl = self.Productos.Control.text().strip()
        NombreProd = self.Productos.NombreProd.text().strip()
        Cantidad = float(self.Productos.Num1.text().strip()) if self.Productos.Num1.text().strip() else 0
        PrecioTT = float(self.Productos.PrecioTT.text().strip()) if self.Productos.PrecioTT.text().strip() else 0
        PrecioUU = float(self.Productos.PrecioUnid.text().strip()) if self.Productos.PrecioUnid.text().strip() else 0

        #Agarrar los datos tipo fecha
        FechaFac = self.Productos.FechaFactu.date().toString('yyyy-MM-dd')
        FechaVenci = self.Productos.FechaVenci.date().toString('yyyy-MM-dd')

        #Agarrar los datos de los ComboBox
        Deposito = int(self.Productos.Deposito.currentText().strip())
        Unidad = self.Productos.Unidad.currentText().strip()
        IVA = float(self.Productos.IVA.currentText().strip('%'))

        # Validar datos
        if not RifEmpr and not NombreProd:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Rif requerido')
            mensaje.exec_()
        elif not Prov:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Nombre del proveedor requerido')
            mensaje.exec_()
        elif not NumFact:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Numero de factura requerido requerido')
            mensaje.exec_()
        elif not NombreProd:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Nombre del producto requerido')
            mensaje.exec_()
        elif Cantidad > 0:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('La cantidad debe ser mayor a cero')
            mensaje.exec_()
        elif PrecioTT > 0:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('El precio total debe ser mayor a cero')
            mensaje.exec_()
        else:
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor()
                query = "SELECT Inicial, PrecioUU FROM inventario WHERE NomProd = %s ORDER BY IdProducto DESC LIMIT 1;"
                try:
                    cursor.execute(query, (NombreProd,))
                    result = cursor.fetchone()
                    if result is not None:
                        # Si ya existe un registro con el mismo nombre de producto, que sume la cantidad vieja con la nueva
                        CantidadInicial = result[0]
                        prev_precio_uu = result[1]
                        TotalCantidad = float(CantidadInicial) + Cantidad
                        new_precio_uu = PrecioTT/Cantidad
                        PrecioProm = float((prev_precio_uu + new_precio_uu)/ 2)
                        query = '''
                            INSERT INTO inventario (Codigo,Rif,NomEmpr,NomProd,Inicial,Deposito,Unidad,PrecioTT,IVA,PrecioUU,
                            FechaFact,FechaVenci,Ingreso,TotalCantidad,PrecioProm,Stock,NumFact,NumControl)
                            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        '''
                        cursor.execute(query, (Code,RifEmpr,Prov,NombreProd,CantidadInicial,Deposito,
                                                Unidad,PrecioTT,IVA,PrecioUU,FechaFac,FechaVenci,Cantidad,TotalCantidad,PrecioProm,0,
                                                NumFact,NumControl))
                        mensaje.setIcon(QMessageBox.Information)
                        mensaje.setText('Informacion Actualizada')
                        mensaje.exec_()

                        self.Productos.Proveedor.setEnabled(False)
                        self.Productos.CodeProduct.setEnabled(False)
                        self.Productos.Factura.setEnabled(False)
                        self.Productos.Control.setEnabled(False)
                        self.Productos.NombreProd.setEnabled(False)
                        self.Productos.Num1.setEnabled(False)
                        self.Productos.PrecioTT.setEnabled(False)
                        self.Productos.FechaFactu.setEnabled(False)
                        self.Productos.FechaVenci.setEnabled(False)
                        self.Productos.Deposito.setEnabled(False)
                        self.Productos.Unidad.setEnabled(False)
                        self.Productos.IVA.setEnabled(False)
                    else:
                        # Si no existe, insertar un nuevo registro
                        query = '''
                            INSERT INTO inventario (Codigo,Rif,NomEmpr,NomProd,Inicial,Deposito,Unidad,PrecioTT,IVA,PrecioUU,
                            FechaFact,FechaVenci,Ingreso,TotalCantidad,PrecioProm,NumFact,NumControl)
                            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        '''
                        cursor.execute(query, (Code,RifEmpr,Prov,NombreProd,Cantidad,Deposito,
                                                Unidad,PrecioTT,IVA,PrecioUU,FechaFac,FechaVenci,0,0,0,
                                                NumFact,NumControl))
                        connection.commit()
                        cursor.close()
                        connection.close()
                        mensaje.setIcon(QMessageBox.Information)
                        mensaje.setText('Producto añadido correctamente')
                        mensaje.exec_()

                        self.Productos.Proveedor.setEnabled(False)
                        self.Productos.CodeProduct.setEnabled(False)
                        self.Productos.Factura.setEnabled(False)
                        self.Productos.Control.setEnabled(False)
                        self.Productos.NombreProd.setEnabled(False)
                        self.Productos.Num1.setEnabled(False)
                        self.Productos.PrecioTT.setEnabled(False)
                        self.Productos.FechaFactu.setEnabled(False)
                        self.Productos.FechaVenci.setEnabled(False)
                        self.Productos.Deposito.setEnabled(False)
                        self.Productos.Unidad.setEnabled(False)
                        self.Productos.IVA.setEnabled(False)
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error al añadir la informacion: {e}')
                    mensaje.exec_()
                connection.close()
            else:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()
    
    def add_prov(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        #Agarrar los datos de los cuadros de texto
        Rif = self.Prov.Rif.text().strip()
        NombreProv = self.Prov.Empresa.text().strip()
        TlfnoEmp = self.Prov.TlfnoEmpre.text().strip()
        Email = self.Prov.EmailEmpre.text().strip()
        NombreVen = self.Prov.NameVen.text().strip()
        TlfnoVen = self.Prov.TlfnoVende.text().strip()

        # Validar datos
        if not Rif:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('El campo Rif del proveedor no puede estar vacío')
            mensaje.exec_()
        elif not NombreProv:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('El campo Nombre Empresa no puede estar vacío')
            mensaje.exec_()
        elif not TlfnoEmp:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('El campo Telefono Empresa no puede estar vacío')
            mensaje.exec_()
        elif not NombreVen:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('El campo Nombre Vendedor no puede estar vacío')
            mensaje.exec_()
        elif not TlfnoVen:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('El campo Telefono Vendedor no puede estar vacío')
            mensaje.exec_()
        else:
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor()
                query = "SELECT COUNT(*) FROM proveedores WHERE Rif = %s"
                try:
                    cursor.execute(query, (Rif,))  # Pasa el valor Rif como parámetro
                    result = cursor.fetchone()[0]
                    if result > 0:  # Verifica si el conteo es mayor que 0
                        mensaje.setIcon(QMessageBox.Warning)
                        mensaje.setText('El proveedor ya está registrado')
                    else:
                        query = '''
                        INSERT INTO proveedores (Rif,NomEmpr,Tlfno,Email,NombreVen,TlfnoVen)
                        VALUES(%s,%s,%s,%s,%s,%s);
                        '''
                        cursor.execute(query, (Rif,NombreProv,TlfnoEmp,Email,NombreVen,TlfnoVen))
                        connection.commit()
                        cursor.close()
                        connection.close()
                        mensaje.setIcon(QMessageBox.Information)
                        mensaje.setText('Proveedor registrado correctamente')
                        mensaje.exec_()

                        #Evita que los cuadros de texto sean editables una vez registrado el proveedor
                        self.Prov.Rif.setEnabled(False)
                        self.Prov.Empresa.setEnabled(False)
                        self.Prov.TlfnoEmpre.setEnabled(False)
                        self.Prov.EmailEmpre.setEnabled(False)
                        self.Prov.NameVen.setEnabled(False)
                        self.Prov.TlfnoVende.setEnabled(False)
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error al añadir el proveedor: {e}')
                    mensaje.exec_()
                connection.close()
            else:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()

    def registrarempl(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        #Agarrar los datos de los cuadros de texto
        Name1 = self.Emple.NameEmpl1.text().strip()
        Name2 = self.Emple.NameEmpl2.text().strip()
        Last1 = self.Emple.ApeEmpl1.text().strip()
        Last2 = self.Emple.ApeEmpl2.text().strip()
        Cedula = self.Emple.Cedu.text().strip()
        Tlfno = self.Emple.CellNumber.text().strip()
        Salario = self.Emple.Salario.text().strip()

        #Agarrar los datos tipo fecha
        Ingreso = self.Emple.FechaIngr.date().toString('yyyy-MM-dd')

        #Agarrar los datos de los ComboBox
        Cargo = self.Emple.Cargo.currentText().strip()

        # Validar datos
        if not Name1:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Falta el primer nombre')
            mensaje.exec_()
        elif not Last1 and Cedula:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Falta el primer apellido')
            mensaje.exec_()
        elif not Cedula:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Cedula requerida')
            mensaje.exec_()
        else:
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor()
                query = "SELECT COUNT(*) FROM empleados WHERE Cedula = %s"
                try:
                    cursor.execute(query, (Cedula,))
                    result = cursor.fetchone()[0]
                    if result > 0:
                        mensaje.setIcon(QMessageBox.Warning)
                        mensaje.setText('El empleado ya está registrado')
                    else:
                        query = '''
                        INSERT INTO empleados (PrimerNombre, SegunNombre, PrimerApe, SegunApe, Cedula, Tlfno,
                        Cargo, FechaIngre, SalarioBase)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s);
                        '''
                        cursor.execute(query, (Name1, Name2, Last1, Last2, Cedula, Tlfno, Cargo, Ingreso, Salario))
                        connection.commit()
                        cursor.close()
                        connection.close()
                        mensaje.setIcon(QMessageBox.Information)
                        mensaje.setText('Empleado registrado correctamente')
                        mensaje.exec_()

                        self.Emple.NameEmpl1.setEnabled(False)
                        self.Emple.NameEmpl2.setEnabled(False)
                        self.Emple.ApeEmpl1.setEnabled(False)
                        self.Emple.ApeEmpl2.setEnabled(False)
                        self.Emple.Cedu.setEnabled(False)
                        self.Emple.CellNumber.setEnabled(False)
                        self.Emple.Salario.setEnabled(False)
                        self.Emple.FechaIngr.setEnabled(False)
                        self.Emple.Cargo.setEnabled(False)
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error al registrar empleado: {e}')
                    mensaje.exec_()
                connection.close()
            else:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()

    def searchEmpl(self):
        NameEmpl = self.Datos.SearchName.text().strip()
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        if len(NameEmpl):
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor() #Busqueda de la informacion en la tabla
                query = "SELECT PrimerNombre, SalarioBase, Cedula, Cargo FROM empleados WHERE PrimerNombre = %s"
                try:
                    cursor.execute(query, (NameEmpl,))
                    result = cursor.fetchone()
                    if result is not None:
                        self.Datos.NombreEmp.setText(result[0])
                        self.Datos.SalarioBase.setText(str(result[1]))
                        self.Datos.Cedu.setText(result[2])
                        self.Datos.Cargo.setText(result[3])
                        self.Datos.SearchName.clear()
                        mensaje.setIcon(QMessageBox.Information)
                        mensaje.setText('Empleado encontrado')
                        mensaje.exec_()
                    else:
                        mensaje.setIcon(QMessageBox.Warning)
                        mensaje.setText('Empleado no encontrado')
                        mensaje.exec_()
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error al buscar el empleado: {e}')
                    mensaje.exec_()
                connection.close()
            else:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()

    def add_datosnomina(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        informacion = extraer_informacion()

        DiaNomina = QDate.currentDate().toString('yyyy-MM-dd') #Agarra directamente la fecha de la computadora

        #Agarrar los datos de los cuadros de texto
        NameEmpl = self.Datos.NombreEmp.text().strip()
        Salario = int(self.Datos.SalarioBase.text().strip())
        Cedula = self.Datos.Cedu.text().strip()
        Cargo = self.Datos.Cargo.text().strip()
        DiasTrab = int(self.Datos.Dias.text().strip())
        DiasFeria = int(self.Datos.Feriados.text().strip()) if self.Datos.Feriados.text().strip() else 0
        Domingos = int(self.Datos.Domingos.text().strip()) if self.Datos.Domingos.text().strip() else 0
        Libres = int(self.Datos.Libres.text().strip()) if self.Datos.Libres.text().strip() else 0
        ExtraDiurnas = int(self.Datos.HExDiurnas.text().strip()) if self.Datos.HExDiurnas.text().strip() else 0
        ExtraNocturnas = int(self.Datos.HExNocturnas.text().strip()) if self.Datos.HExNocturnas.text().strip() else 0
        BonoNoct = int(self.Datos.BonoNoct.text().strip()) if self.Datos.BonoNoct.text().strip() else 0
        DiasReposo = int(self.Datos.Reposo.text().strip()) if self.Datos.Reposo.text().strip() else 0
        CestaTicket = int(self.Datos.CT.text().strip())
        islr = float(self.Datos.ISRL.text().strip()) if self.Datos.ISRL.text().strip() else 0
        Faltante = float(self.Datos.Faltan.text().strip()) if self.Datos.Faltan.text().strip() else 0
        Prestamo = float(self.Datos.Presta.text().strip()) if self.Datos.Presta.text().strip() else 0
        Consumo = float(self.Datos.Consu.text().strip()) if self.Datos.Consu.text().strip() else 0
        
        #Agarrar los datos de los ComboBox
        Periodo = self.Datos.Periodo.currentText()
        Lunes = int(self.Datos.Lunes.currentText())
        DiasDesc = int(self.Datos.Descanso.currentText())

        SaD = (Salario)/30 #Salario del día
        Sb = (SaD)*DiasTrab #Salario del día por días trabajados
        DeD = (SaD)*(DiasDesc)#Salario del día por días de descanso
        Fer = (SaD)*(DiasFeria*1.5) #Salario del día por días feriados trabajados
        Do = (SaD)*(Domingos*1.5) #Salario del día por domingos trabajados
        Lib = (SaD)*(Libres*1.5) #Salario del día por días libres trabajados
        Hed = ((SaD)/8)*(ExtraDiurnas*1.5) #Salario del día más las horas extra diurnas
        Hen = ((SaD)/8)*(ExtraDiurnas*1.5*1.3) #Salario del día más las horas extra nocturnas
        BoN = ((SaD)/8)*(BonoNoct*0.3) #Salario del día más bono nocturno
        DiasRestanRepo = DiasReposo-3

        if DiasReposo <= 3 and DiasReposo > 0:
            Rep = (SaD)*DiasReposo
        elif DiasReposo > 3:
            Rep = ((SaD)*3)+(DiasRestanRepo*0.3533)
        else:
            Rep = 0

        Deveng = Sb+DeD+Fer+Do+Lib+Hed+Hen+BoN+Rep
        TDeveng = round(Deveng * informacion, 2) #Utilizar la información extraída
        Porcenislr = islr/100
        #Deducciones
        sso = round((((300*0.04)*Lunes)/informacion), 2)
        rpe = round((((300*0.005)*Lunes)/informacion), 2)
        faov = round((Deveng*0.01),2)
        Islr = round((Deveng*Porcenislr),2)

        DiasCT = DiasTrab+DiasDesc+DiasReposo
        TotalCT = round((DiasCT/30)*CestaTicket, 2)
        TDes = sso+rpe+faov+Islr
        TSalCT = (Deveng+TotalCT)-TDes

        TotalPagar = (TSalCT)-Prestamo-Faltante-Consumo

        # Validar datos
        if not NameEmpl:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Falta el nombre del empleado')
            mensaje.exec_()
        elif not Salario:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Falta el salario del empleado')
            mensaje.exec_()
        elif not Cedula:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Falta la cedula del empleado')
            mensaje.exec_()
        elif not DiasTrab:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Faltan los dias trabajados del empleado')
            mensaje.exec_()
        elif not Domingos:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Faltan los domingos trabajados por el empleado')
            mensaje.exec_()
        elif not CestaTicket:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Falta el monto de cestaticket del empleado')
            mensaje.exec_()
        else:
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor()
                query = "SELECT COUNT(*) FROM nomina WHERE NombreEmpl = %s"
                try:
                    cursor.execute(query, (NameEmpl,))
                    result = cursor.fetchone()[0]
                    query = '''
                    INSERT INTO nomina (FechaNomina,NombreEmpl,Cedula,Cargo,SalarioBase,Periodo,DiasTrab,Salario,Descanso,
                    Des,Feriados,Fer,Domingos,Domin,Libres,Lib,HorasDiurnas,Hed,HorasNocturnas,Hen,BonoNoctu,BoN,Reposo,Rep,
                    Devengado,TotalDeveng,SSO,RPE,FAOV,ISLR,MontoCT,DiasCT,TotalCT,TotalDesc,TotalSalarioCT,Descuento,Faltante,Consumo,TotalPagar)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
                    '''
                    cursor.execute(query, (DiaNomina,NameEmpl,Cedula,Cargo,Salario,Periodo,DiasTrab,Sb,DiasDesc,DeD,DiasFeria,Fer,Domingos,
                                    Do,Libres,Lib,ExtraDiurnas,Hed,ExtraNocturnas,Hen,BonoNoct,BoN,DiasReposo,Rep,Deveng,
                                    TDeveng,sso,rpe,faov,Islr,CestaTicket,DiasCT,TotalCT,TDes,TSalCT,Prestamo,Faltante,Consumo,TotalPagar))
                    connection.commit()
                    cursor.close()
                    connection.close()
                    mensaje.setIcon(QMessageBox.Information)
                    mensaje.setText('Nomina del empleado registrado exitosamente')
                    mensaje.exec_()

                    self.Datos.SearchName.setEnabled(False)
                    self.Datos.SalarioBase.setEnabled(False)
                    self.Datos.Dias.setEnabled(False)
                    self.Datos.Feriados.setEnabled(False)
                    self.Datos.Domingos.setEnabled(False)
                    self.Datos.Libres.setEnabled(False)
                    self.Datos.HExDiurnas.setEnabled(False)
                    self.Datos.HExNocturnas.setEnabled(False)
                    self.Datos.BonoNoct.setEnabled(False)
                    self.Datos.Reposo.setEnabled(False)
                    self.Datos.CT.setEnabled(False)
                    self.Datos.ISRL.setEnabled(False)
                    self.Datos.Faltan.setEnabled(False)
                    self.Datos.Presta.setEnabled(False)
                    self.Datos.Consu.setEnabled(False)
                    self.Datos.Periodo.setEnabled(False)
                    self.Datos.Lunes.setEnabled(False)
                    self.Datos.Descanso.setEnabled(False)
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error añadir los datos a la nómina: {e}')
                    mensaje.exec_()
                connection.close()
            else:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()
        
    def actualizar_tablas1(self):
        self.mostrarInven()
        self.mostrarNomina()
        self.mostrarAuditoria()
        self.ventas()
        self.montos()
        self.mostrarmetodo()
        self.mostrar_precio()
        self.timer.singleShot(1000, self.actualizar_tablas1)

    def actualizar_hora(self):
        self.Venta.Hora.setTime(QTime.currentTime())

    def ventas(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM venta
            '''
            cursor.execute(query,)
            results = cursor.fetchall() #Obtener todos los resultados de la consulta
            people = []
            for result in results:
                Nombre = result[1]
                Cantidad = result[2]
                Precio = result[3]
                
                people.append({"Descripcion":Nombre,"Cantidad":Cantidad,"Precio":Precio})
            rows = 0
            self.Venta.tableWidget.setRowCount(0)
            for item in people:
                self.Venta.tableWidget.insertRow(rows)
                rows+=1
                for j, value in enumerate(item.values()):
                    self.Venta.tableWidget.setItem(rows-1, j, QTableWidgetItem(str(value)))
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()
    
    def borrarItem(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
    
        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM venta
            '''
            cursor.execute(query,)
            results = cursor.fetchall() #Obtener todos los resultados de la consulta
            try:
                if connect is not None:
                    query='''
                    DELETE FROM venta
                    WHERE IdVenta IN (
                        SELECT IdVenta
                        FROM (
                            SELECT IdVenta
                            FROM venta
                            ORDER BY IdVenta DESC
                            LIMIT 1
                        ) as subquery
                    )
                    '''
                    cursor.execute(query)
                    connection.commit()
                    cursor.close()
                    connection.close()
                    self.mostrar_precio()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error al borrar el item: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def mostrar_precio(self):
        mensaje = QMessageBox(self)
        mensaje.setWindowTitle('Información')

        informacion = extraer_informacion()
        connection = self.conectarBD()
        if connection is not None:
            try:
                cursor = connection.cursor()
                query = '''
                    SELECT SUM(Precio) as TotalPrecio FROM venta
                '''
                cursor.execute(query,)
                result = cursor.fetchone()
                if result is not None and result[0] is not None:
                    precio_total = result[0] if result[0] is not None else 0
                    iva_rate = 0.16

                    total = precio_total*informacion

                    iva = (precio_total*informacion) * iva_rate
                    sub_total = (precio_total*informacion) - iva
                    total_d = precio_total

                    self.Venta.SubTotal.setText(f"{sub_total:.2f}")
                    self.Venta.iva.setText(f"{iva:.2f}")
                    self.Venta.Pagarbs.setText(f"{total:.2f}")
                    self.Venta.PagarD.setText(f"{total_d:.2f}")
                    connection.commit()
                else:
                    self.Venta.SubTotal.clear()
                    self.Venta.iva.clear()
                    self.Venta.Pagarbs.clear()
                    self.Venta.PagarD.clear()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error al mostrar el precio: {str(e)}')
                mensaje.exec_()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('No se pudo conectar a la base de datos')
            mensaje.exec_()

    def cuartopollocomer(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre, Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (1,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def cuartopollollevar(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (2,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
    
    def mediocomer(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (3,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def mediollevar(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (4,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def seiscomer(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (5,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def seisllevar(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (6,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def pollocomer(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (7,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def pollollevar(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = "SELECT Nombre,Precio FROM menu WHERE Id = %s"
            try:
                cursor.execute(query, (8,))
                result = cursor.fetchone()
                if result is not None and len(result) >= 2:
                    Nombre = result[0]
                    Precio = result[1]
                    Cantidad = 1
                    query = '''
                    INSERT INTO venta (Nombre,Cantidad,Precio) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query,(Nombre,Cantidad,Precio))
                    connection.commit()
                    cursor.close()
                    connection.close()
                else:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText('No se ha encontrado el producto')
                    mensaje.exec_()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error añadir venta: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def montos(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        informacion = float(extraer_informacion())
        connection = self.conectarBD()
        if connection is not None:
            try:
                cursor = connection.cursor()
                query = '''
                    SELECT SUM(Precio) as TotalPrecio FROM venta
                '''
                cursor.execute(query,)
                result = cursor.fetchone()
                if result is not None and result[0] is not None:
                    Monto = float(result[0]) if result[0] is not None else 0
                    subtotal = float((Monto * informacion)*0.16)
                    TotalBs = (Monto * informacion)+subtotal
                    subtotalD = float(Monto*0.03)
                    TotalD = (Monto + subtotalD)

                    self.Pago.BS.setText(f"{TotalBs:.2f}")
                    self.Pago.Divisa.setText(f"{Monto:.2f}")
                    self.Pago.igtf.setText(f"{subtotalD:.2f}")
                    self.Pago.TotalPagar.setText(f"{TotalD:.2f}")
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error al registrar la forma de pago: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    def vuelto(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        Lista = []

        informacion = extraer_informacion()

        if self.Pago.BS.text().strip():
            TotalBs = float(self.Pago.BS.text().strip())
        else:
            TotalBs = 0.0

        if self.Pago.TotalPagar.text().strip():
            TotalD = float(self.Pago.TotalPagar.text().strip())
        else:
            TotalD = 0.0

        currentIndex = self.Pago.listWidget.currentRow()
        item = self.Pago.listWidget.item(currentIndex)
        if item is not None:
            Lista = self.Pago.listWidget.currentItem().text().strip()
        montocliente = float(self.Pago.Monto.text().strip().replace(',','.'))

        if montocliente>0:
            if Lista =='Bs Efectivo':
                self.Pago.FormaPa.setText("Bs Efectivo")
                #Vuelto Bs
                if montocliente > TotalBs:
                    Vuelto = montocliente - TotalBs
                    self.Pago.Vuelto.setText(f"{Vuelto:.2f}")
                elif montocliente < TotalBs:
                    Restobs = TotalBs - montocliente
                    RestanteD = TotalD - (montocliente/informacion)
                    self.Pago.BS.setText(f"{Restobs:.2f}")
                    self.Pago.TotalPagar.setText(f"{RestanteD:.2f}")
                else:
                    self.Pago.Vuelto.setText("0.00")
            elif Lista == 'Tarjeta de Debito':
                self.Pago.FormaPa.setText("TDD")
                #Vuelto Bs
                if montocliente > TotalBs:
                    Vuelto = montocliente - TotalBs
                    self.Pago.Vuelto.setText(f"{Vuelto:.2f}")
                elif montocliente < TotalBs:
                    Restobs = TotalBs - montocliente
                    RestanteD = TotalD - (montocliente/informacion)
                    self.Pago.BS.setText(f"{Restobs:.2f}")
                    self.Pago.TotalPagar.setText(f"{RestanteD:.2f}")
                else:
                    self.Pago.Vuelto.setText("0.00")
            elif Lista == 'Tarjeta de Credito':
                self.Pago.FormaPa.setText("TDC")
                #Vuelto Bs
                if montocliente > TotalBs:
                    Vuelto = montocliente - TotalBs
                    self.Pago.Vuelto.setText(f"{Vuelto:.2f}")
                elif montocliente < TotalBs:
                    Restobs = TotalBs - montocliente
                    RestanteD = TotalD - (montocliente/informacion)
                    self.Pago.BS.setText(f"{Restobs:.2f}")
                    self.Pago.TotalPagar.setText(f"{RestanteD:.2f}")
                else:
                    self.Pago.Vuelto.setText("0.00")
            elif Lista == 'Pago Movil':
                self.Pago.FormaPa.setText("Pago Movil")
                #Vuelto Bs
                if montocliente > TotalBs:
                    Vuelto = montocliente - TotalBs
                    self.Pago.Vuelto.setText(f"{Vuelto:.2f}")
                elif montocliente < TotalBs:
                    Restobs = TotalBs - montocliente
                    RestanteD = TotalD - (montocliente/informacion)
                    self.Pago.BS.setText(f"{Restobs:.2f}")
                    self.Pago.TotalPagar.setText(f"{RestanteD:.2f}")
                else:
                    self.Pago.Vuelto.setText("0.00")
            elif Lista == 'Biopago':
                self.Pago.FormaPa.setText("Biopago")
                #Vuelto Bs
                if montocliente > TotalBs:
                    Vuelto = montocliente - TotalBs
                    self.Pago.Vuelto.setText(f"{Vuelto:.2f}")
                elif montocliente < TotalBs:
                    Restobs = TotalBs - montocliente
                    RestanteD = TotalD - (montocliente/informacion)
                    self.Pago.BS.setText(f"{Restobs:.2f}")
                    self.Pago.TotalPagar.setText(f"{RestanteD:.2f}")
                else:
                    self.Pago.Vuelto.setText("0.00")
            else:
                self.Pago.FormaPa.setText("Divisa")
                #Vuelto Divisa
                if montocliente > TotalD:
                    Vuelto = montocliente - TotalD
                    self.Pago.Vuelto.setText(f"{Vuelto:.2f}")
                elif TotalD > montocliente:
                    RestoD = TotalD - montocliente
                    Restantebs = TotalBs - (montocliente * informacion)
                    self.Pago.TotalPagar.setText(f"{RestoD:.2f}")
                    self.Pago.BS.setText(f"{Restantebs:.2f}")
                else:
                    self.Pago.Vuelto.setText("0.00")
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Monto Invalido')
            mensaje.exec_()

    def registrarpago(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        Metodo = self.Pago.FormaPa.text().strip()
        Monto = self.Pago.Monto.text().strip()
        Referencia = self.Pago.RefePM.text().strip()

        # Validar datos
        if not Metodo and Monto:
            print('Faltan datos requeridos')
        else:
            connection = self.conectarBD()
            if connection is not None:
                cursor = connection.cursor()
                query = '''
                SELECT * FROM metodopago
                '''
                cursor.execute(query)
                result = cursor.fetchone()
                try:
                    query = '''
                    INSERT INTO metodopago (Nombre,Monto,NoReferencia) VALUES(%s,%s,%s);
                    '''
                    cursor.execute(query, (Metodo,Monto,Referencia))
                    connection.commit()
                    cursor.close()
                    connection.close()
                except Exception as e:
                    mensaje.setIcon(QMessageBox.Warning)
                    mensaje.setText(f'Error al registrar la informacion: {e}')
                    mensaje.exec_()
                connection.close()
            else: 
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText('Error al conectar a la base de datos')
                mensaje.exec_()
            connection.close()

    def mostrarmetodo(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM metodopago
            '''
            cursor.execute(query,)
            results = cursor.fetchall() #Obtener todos los resultados de la consulta
            people = []
            for result in results:
                Nombre = result[1]
                Monto = result[2]
                Referencia = result[3]
                
                people.append({"Metodo de Pago":Nombre,"Monto":Monto,"NoReferencia":Referencia})
            rows = 0
            self.Pago.tableWidget.setRowCount(0)
            for item in people:
                self.Pago.tableWidget.insertRow(rows)
                rows+=1
                for j, value in enumerate(item.values()):
                    self.Pago.tableWidget.setItem(rows-1, j, QTableWidgetItem(str(value)))
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def factura(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        fecha = fecha_actual.toString('yyyy-MM-dd')
        hora = hora_actual.toString('HH:mm:ss')
        Nofactura = 0

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            try:
                query = '''
                SELECT * FROM venta
                '''
                cursor.execute(query,)
                results = cursor.fetchall() #Obtener todos los resultados de la consulta
                for result in results:
                    Nombre = result[1]
                    Cantidad = result[2]
                    Precio = result[3]

                    query = '''
                        SELECT SUM(Precio) as TotalPrecio FROM venta
                    '''
                    cursor.execute(query,)
                    result = cursor.fetchone()
                    TotalPrecio = round(result[0],2)

                    buffer = io.BytesIO()
                    try:
                        c = canvas.Canvas(buffer, pagesize=letter)

                        c.drawString(100, 750, f'Factura No. {Nofactura}')
                        c.drawString(100, 730, f'Fecha: {fecha}')
                        c.drawString(100, 710, f'Hora: {hora}')
                        c.drawString(100, 690, f'Nombre: {Nombre}')
                        c.drawString(100, 670, f'Cantidad: {Cantidad}')
                        c.drawString(100, 600, f'Precio: {Precio}')
                        c.drawString(100, 650, f'Total: {TotalPrecio}')

                        c.showPage()
                        c.save()
                        
                        buffer.seek(0)  # Regresar el puntero al principio del buffer

                        #Guardar pdf en el escritorio
                        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                        with open(os.path.join(desktop_path, f"Factura No. {Nofactura}.pdf"), 'wb') as f:
                            f.write(buffer.getvalue())
                    except Exception as e:
                        mensaje.setIcon(QMessageBox.Warning)
                        mensaje.setText(f'Error al crear el archivo PDF: {e}')
                        mensaje.exec_()
                    finally:
                        buffer.close() #cerrar el buffer

                    #Guardar la informacion de la factura en la base de datos
                    query = '''
                        INSERT INTO facturacion (NumeroFactura,Fecha,Hora,Nombre,Cantidad,Precio) VALUES(%s,%s,%s,%s,%s,%s);
                        '''
                    cursor.execute(query, (Nofactura,fecha,hora,Nombre,Cantidad,Precio))
                    connection.commit()
                    
                    Nofactura += 1

                    #Vaciar la tabla venta
                    query = "TRUNCATE TABLE venta"
                    cursor.execute(query)
                    connection.commit()
                
                    #Vaciar la tabla metodopago
                    query = "TRUNCATE TABLE metodopago"
                    cursor.execute(query)
                    connection.commit()

                cursor.close()
                connection.close()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error al generar la factura: {e}')
                mensaje.exec_()
            connection.close()

        #Regresar a la ventana Venta
        self.widget.setCurrentWidget(self.Venta)

    def realizar_auditoria(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            try:
                # Obtener datos de la tabla facturacion
                query_facturacion = '''
                    SELECT Nombre, Cantidad FROM facturacion
                '''
                cursor.execute(query_facturacion)
                facturacion_data = cursor.fetchall()

                for facturacion_row in facturacion_data:
                    facturacionProduc = facturacion_row[0]
                    facturacionCantidad = facturacion_row[1]

                # Obtener datos de la tabla inventario
                query_inventario = '''
                    SELECT NomProd, TotalCantidad, PrecioUU FROM inventario WHERE NomProd = %s
                '''
                cursor.execute(query_inventario, (facturacionProduc,))
                inventario_row = cursor.fetchone()

                # Comparar nombres de producto y actualizar cantidades

                if inventario_row:
                    inventarioProduc = inventario_row[0]
                    inventarioCantidad = inventario_row[1]
                    inventarioPrecio = inventario_row[2]
                else:
                    inventarioProduc = facturacionProduc
                    inventarioCantidad = 0
                    inventarioPrecio = 0

                stock = inventarioCantidad - facturacionCantidad
                #Registrar la información en la tabla auditoria
                query = '''
                    INSERT INTO auditoria(NombreProducto,CantidadInventario,CantidadVenta,PrecioUnitCompra,Existencia)
                    VALUES (%s, %s, %s, %s, %s);
                '''
                cursor.execute(query, (facturacionProduc, inventarioCantidad, facturacionCantidad, inventarioPrecio, stock))
                connection.commit()
            except Exception as e:
                mensaje.setIcon(QMessageBox.Warning)
                mensaje.setText(f'Error al realizar la auditoría: {e}')
                mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()

    #Accion de cambiar entre ventanas al presionar un boton
    def conexiones(self):
        #Ventana Inicio Sesion
        self.Inicio.Entrar.clicked.connect(self.login)
        self.Inicio.Exit.clicked.connect(self.close)
        self.Inicio.Password.clear()

        #Ventana Menu
        self.Menu.Inv.clicked.connect(self.cambio_Invent)
        self.Menu.Exit.clicked.connect(self.cambio_Inicio)
        self.Menu.Nomina.clicked.connect(self.cambio_Nom)
        self.Menu.Venta.clicked.connect(self.cambio_Venta)
        self.Menu.Auditoria.clicked.connect(self.cambio_audit)

        #Ventana Inventario
        self.Invent.Exit.clicked.connect(self.cambio_Menu)
        self.Invent.AddEdit.clicked.connect(self.cambio_Product)
        self.Invent.Download.clicked.connect(self.imprimirinvent)

        #Ventana Productos
        self.Productos.Search.clicked.connect(self.buscar_product)
        self.Productos.SearchRif.clicked.connect(self.buscar_rif)
        self.Productos.Exit.clicked.connect(self.cambio_Invent)
        self.Productos.Exit.clicked.connect(self.limpiar_productos)
        self.Productos.AddProv.clicked.connect(self.cambio_Prov)
        self.Productos.AddProd.clicked.connect(self.add_info)
        self.Productos.Limpiar.clicked.connect(self.limpiar_productos)
        self.Productos.IVAsi.toggled.connect(self.calculate_iva)
        self.Productos.FechaFactu.setDate(fecha_actual)
        self.Productos.FechaVenci.setDate(fecha_actual)

        #Ventana Proveedores
        self.Prov.Exit.clicked.connect(self.cambio_Product)
        self.Prov.Exit.clicked.connect(self.limpiar_prov)
        self.Prov.AddProv.clicked.connect(self.add_prov)
        self.Prov.Limpiar.clicked.connect(self.limpiar_prov)

        #Ventana Nomina
        self.Nomina.Exit.clicked.connect(self.cambio_Menu)
        self.Nomina.AddEmpl.clicked.connect(self.cambio_emple)
        self.Nomina.Datos.clicked.connect(self.cambio_datos)
        self.Nomina.Download.clicked.connect(self.imprimirnomina)

        #Ventana Empleados
        self.Emple.Exit.clicked.connect(self.cambio_Nom)
        self.Emple.Exit.clicked.connect(self.limpiar_empleados)
        self.Emple.Limpiar.clicked.connect(self.limpiar_empleados)
        self.Emple.AddEmpl.clicked.connect(self.registrarempl)
        self.Emple.FechaIngr.setDate(fecha_actual)

        #Ventana Datos Nomina
        self.Datos.Exit.clicked.connect(self.cambio_Nom)
        self.Datos.Exit.clicked.connect(self.limpiar_datosnomi)
        self.Datos.Add.clicked.connect(self.add_datosnomina)
        self.Datos.Limpiar.clicked.connect(self.limpiar_datosnomi)
        self.Datos.Search.clicked.connect(self.searchEmpl)

        #Ventana Venta
        self.Venta.Exit.clicked.connect(self.cambio_Menu)
        self.Venta.Exit.clicked.connect(self.limpiar_venta)
        self.Venta.Tasa.setText(str(self.conexionbcv()))
        self.Venta.Fecha.setDate(fecha_actual)
        self.Venta.Hora.setTime(hora_actual)
        self.Venta.DeleteItem.clicked.connect(self.borrarItem)
        self.Venta.cuartocomer.clicked.connect(self.cuartopollocomer)
        self.Venta.cuartollevar.clicked.connect(self.cuartopollollevar)
        self.Venta.mediocomer.clicked.connect(self.mediocomer)
        self.Venta.mediollevar.clicked.connect(self.mediollevar)
        self.Venta.sixcomer.clicked.connect(self.seiscomer)
        self.Venta.sixllevar.clicked.connect(self.seisllevar)
        self.Venta.pollocomer.clicked.connect(self.pollocomer)
        self.Venta.pollollevar.clicked.connect(self.pollollevar)
        self.Venta.cuartocomer.clicked.connect(self.mostrar_precio)
        self.Venta.cuartollevar.clicked.connect(self.mostrar_precio)
        self.Venta.mediocomer.clicked.connect(self.mostrar_precio)
        self.Venta.mediollevar.clicked.connect(self.mostrar_precio)
        self.Venta.sixcomer.clicked.connect(self.mostrar_precio)
        self.Venta.sixllevar.clicked.connect(self.mostrar_precio)
        self.Venta.pollocomer.clicked.connect(self.mostrar_precio)
        self.Venta.pollollevar.clicked.connect(self.mostrar_precio)
        self.Venta.Pagar.clicked.connect(self.cambio_pagar)

        #Ventana Forma de Pago
        self.Pago.Monto.textChanged.connect(lambda: QTimer.singleShot(2500, lambda: self.vuelto()))
        self.Pago.Registrar.clicked.connect(self.registrarpago)
        self.Pago.Totalizar.clicked.connect(self.factura)
        self.Pago.Totalizar.clicked.connect(self.realizar_auditoria)

        #Ventana Auditoria
        self.Audi.Exit.clicked.connect(self.cambio_Menu)
        self.Audi.Download.clicked.connect(self.imprimirauditoria)

    #Eventos de cambio entre ventanas
    def cambio_Menu(self):
        self.widget.setCurrentWidget(self.Menu)
    
    def cambio_Inicio(self):
        self.widget.setCurrentWidget(self.Inicio)
    
    def cambio_Invent(self):
        self.widget.setCurrentWidget(self.Invent)

    def cambio_Product(self):
        self.widget.setCurrentWidget(self.Productos)
    
    def cambio_Nom(self):
        self.widget.setCurrentWidget(self.Nomina)

    def cambio_Prov(self):
        self.widget.setCurrentWidget(self.Prov)

    def cambio_Venta(self):
        self.widget.setCurrentWidget(self.Venta)

    def cambio_emple(self):
        self.widget.setCurrentWidget(self.Emple)
    
    def cambio_datos(self):
        self.widget.setCurrentWidget(self.Datos)
    
    def cambio_audit(self):
        self.widget.setCurrentWidget(self.Audi)
    
    def cambio_pagar(self):
        self.widget.setCurrentWidget(self.Pago)

    def conexionbcv(self):
        informacion = round(extraer_informacion(),2)
        return informacion

    def mostrarInven(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM inventario
            '''
            cursor.execute(query,)
            results = cursor.fetchall() #Obtener todos los resultados de la consulta
            people = []
            for result in results:
                IdProductos = result[0]
                Codigo = result[1]
                Rif = result[2]
                NombreEmp = result[3]
                NombreProdu = result[4]
                Inicial = result[5]
                Deposito = result[6]
                Unidad = result[7]
                PrecioTT = result[8]
                iva = result[9]
                PrecioUU = result[10]
                FechaFact = result[11]
                FechaVenci = result[12]
                Ingreso = result[13]
                TotalCantidad = result[14]
                PrecioProm = result[15]
                NoFactura = result[16]
                NoControl = result[17]
                people.append({"Id":IdProductos,"Codigo":Codigo,"Rif":Rif,"Nombre Empresa":NombreEmp,
                            "Nombre Producto":NombreProdu,"Inicial":Inicial,"Deposito":Deposito,"Unidad":Unidad,
                            "Precio Total":PrecioTT,"IVA":iva,"Precio Unitario":PrecioUU,"Fecha Factura":FechaFact,
                            "Fecha Vencimiento":FechaVenci,"Ingreso":Ingreso,"Total Cantidad":TotalCantidad,
                            "Precio Promedio":PrecioProm,"No Factura":NoFactura,"No Control":NoControl})

            rows = 0
            self.Invent.tableWidget.setRowCount(0)
            for item in people:
                self.Invent.tableWidget.insertRow(rows)
                rows+=1
                for j, value in enumerate(item.values()):
                    self.Invent.tableWidget.setItem(rows-1, j, QTableWidgetItem(str(value)))
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def mostrarNomina(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM nomina
            '''
            cursor.execute(query,)
            results = cursor.fetchall() #Obtener todos los resultados de la consulta
            people = []
            for result in results:
                IDNomina = result[0]
                FechaNomina = result[1]
                NombreEmpl = result[2]
                Cedula = result[3]
                Cargo = result[4]
                SalarioBase = result[5]
                Periodo	= result[6]
                DiasTrab= result[7]
                Salario= result[8]
                Descanso= result[9]
                Des= result[10]
                Feriados= result[11]
                Fer= result[12]
                Domingos= result[13]
                Domin= result[14]
                Libres= result[15]
                Lib= result[16]
                HorasDiurnas= result[17]
                Hed= result[18]
                HorasNocturnas= result[19]
                Hen= result[20]
                BonoNoctu= result[21]
                BoN= result[22]
                Reposo= result[23]
                Rep= result[24]
                Devengado= result[25]
                TotalDeveng= result[26]
                sso= result[27]
                rpe= result[28]
                faov= result[29]
                isrl= result[30]
                MontoCT= result[31]
                DiasCT= result[32]
                TotalCT= result[33]
                TotalDesc= result[34]
                TotalSalarioCT= result[35]
                Descuento= result[36]
                Faltante= result[37]
                Consumo= result[38]
                TotalPagar= result[39]

                people.append({"ID":IDNomina,"Fecha Nómina":FechaNomina, "Nombre Empleado":NombreEmpl,"CI":Cedula,
                                "Cargo":Cargo,"Salario Base":SalarioBase,"Periodo":Periodo,"Días Trabajados":DiasTrab,
                                "Salario":Salario,"Días Descanso":Descanso,"Descanso":Des,"Días Feriados":Feriados,
                                "Feriados":Fer,"Días Domingos":Domingos,"Domingos":Domin,"Libres Trabajados":Libres,
                                "Libres":Lib,"Horas Diurnas":HorasDiurnas,"Extra Diurnas":Hed,"Horas Nocturnas":HorasNocturnas,
                                "Extra Nocturnas":Hen,"Bono Nocturno":BonoNoctu,"Bono":BoN,"Días Reposo":Reposo,"Reposo":Rep,
                                "Devengado":Devengado,"Total Devengado":TotalDeveng,"SSO":sso,"RPE":rpe,"FAOV":faov,"ISRL":isrl,
                                "Monto CT":MontoCT,"Días CT":DiasCT,"Total CT":TotalCT,"Total Deducciones":TotalDesc,"TotalSalarioCT":TotalSalarioCT,
                                "Descuento":Descuento,"Consumo":Consumo,"Faltante":Faltante,"Total Pagar":TotalPagar})

            rows = 0
            self.Nomina.tableWidget.setRowCount(0)
            for item in people:
                self.Nomina.tableWidget.insertRow(rows)
                rows+=1
                for j, value in enumerate(item.values()):
                    self.Nomina.tableWidget.setItem(rows-1, j, QTableWidgetItem(str(value)))
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def mostrarAuditoria(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM auditoria
            '''
            cursor.execute(query,)
            results = cursor.fetchall() #Obtener todos los resultados de la consulta
            people = []
            for result in results:
                IdProductos = result[0]
                FactuProduc = result[1]
                InventCanti = result[2]
                FactuCanti= result[3]
                InventPrecio = result[4]
                Stock = result[5]

                people.append({"IdProducto":IdProductos,"NompreProducto":FactuProduc,"Cantidad en Inventario":InventCanti,"Cantidad Vendido":FactuCanti,
                            "Precio Unitario":InventPrecio,"Existencia":Stock})

            rows = 0
            self.Audi.tableWidget.setRowCount(0)
            for item in people:
                self.Audi.tableWidget.insertRow(rows)
                rows+=1
                for j, value in enumerate(item.values()):
                    self.Audi.tableWidget.setItem(rows-1, j, QTableWidgetItem(str(value)))
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def imprimirinvent(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')

        # Crea un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM inventario
            '''
            cursor.execute(query,)
            headers = [column[0] for column in cursor.description] #Obtiene los encabezados y los datos de la consulta
            data = cursor.fetchall()

            #Escribe los encabezados en la primera fila de la hoja de cálculo
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            #Escribe los datos en las filas siguientes
            for row_num, row_data in enumerate(data, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_data)

            #Ajusta el ancho de las columnas en función del contenido de las celdas
            for col_num in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_num)].auto_size = True

            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            ws.protection.enable() #Protege la hoja de cálculo contra la modificación
            wb.save(os.path.join(desktop_path, 'Inventario.xlsx')) #Guarda el libro de Excel
            mensaje.setIcon(QMessageBox.Information)
            mensaje.setText('Datos guardados con Éxito')
            mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def imprimirnomina(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        
        # Crea un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM nomina
            '''
            cursor.execute(query,)
            headers = [column[0] for column in cursor.description] #Obtiene los encabezados y los datos de la consulta
            data = cursor.fetchall()

            #Escribe los encabezados en la primera fila de la hoja de cálculo
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            #Escribe los datos en las filas siguientes
            for row_num, row_data in enumerate(data, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_data)

            #Ajusta el ancho de las columnas en función del contenido de las celdas
            for col_num in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_num)].auto_size = True

            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            ws.protection.enable() #Protege la hoja de cálculo contra la modificación
            wb.save(os.path.join(desktop_path, 'Nomina.xlsx')) #Guarda el libro de Excel
            mensaje.setIcon(QMessageBox.Information)
            mensaje.setText('Datos guardados con Éxito')
            mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def imprimirauditoria(self):
        mensaje = QMessageBox(self) #Cuadros de informacion
        mensaje.setWindowTitle('Información')
        
        # Crea un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active

        connection = self.conectarBD()
        if connection is not None:
            cursor = connection.cursor()
            query = '''
            SELECT * FROM auditoria
            '''
            cursor.execute(query,)
            headers = [column[0] for column in cursor.description] #Obtiene los encabezados y los datos de la consulta
            data = cursor.fetchall()

            #Escribe los encabezados en la primera fila de la hoja de cálculo
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            #Escribe los datos en las filas siguientes
            for row_num, row_data in enumerate(data, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_data)

            #Ajusta el ancho de las columnas en función del contenido de las celdas
            for col_num in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_num)].auto_size = True

            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            ws.protection.enable() #Protege la hoja de cálculo contra la modificación
            wb.save(os.path.join(desktop_path, 'Auditoria.xlsx')) #Guarda el libro de Excel
            mensaje.setIcon(QMessageBox.Information)
            mensaje.setText('Datos guardados con Éxito')
            mensaje.exec_()
            connection.close()
        else:
            mensaje.setIcon(QMessageBox.Warning)
            mensaje.setText('Error al conectar a la base de datos')
            mensaje.exec_()
        connection.close()

    def limpiar_productos(self): #Limpiar cuadros de texto de la ventana Productos
        self.Productos.BuscarProd.clear()
        self.Productos.Rif.clear()
        self.Productos.Proveedor.clear()
        self.Productos.BuscarProd.clear()
        self.Productos.CodeProduct.clear()
        self.Productos.Factura.clear()
        self.Productos.Control.clear()
        self.Productos.NombreProd.clear()
        self.Productos.Num1.clear()
        self.Productos.PrecioTT.clear()
        self.Productos.PrecioUnid.clear()
        self.Productos.PrecioIVA.clear()
        self.Productos.Rif.setEnabled(True)
        self.Productos.Proveedor.setEnabled(True)
        self.Productos.CodeProduct.setEnabled(True)
        self.Productos.Factura.setEnabled(True)
        self.Productos.Control.setEnabled(True)
        self.Productos.NombreProd.setEnabled(True)
        self.Productos.Num1.setEnabled(True)
        self.Productos.PrecioTT.setEnabled(True)
        self.Productos.FechaFactu.setEnabled(True)
        self.Productos.FechaVenci.setEnabled(True)
        self.Productos.Deposito.setEnabled(True)
        self.Productos.Unidad.setEnabled(True)
        self.Productos.IVA.setEnabled(True)

    def limpiar_empleados(self): #Limpiar cuadros de texto de la ventana Empleados
        self.Emple.NameEmpl1.clear()
        self.Emple.NameEmpl2.clear()
        self.Emple.ApeEmpl1.clear()
        self.Emple.ApeEmpl2.clear()
        self.Emple.Cedu.clear()
        self.Emple.CellNumber.clear()
        self.Emple.Salario.clear()
        self.Emple.NameEmpl1.setEnabled(True)
        self.Emple.NameEmpl2.setEnabled(True)
        self.Emple.ApeEmpl1.setEnabled(True)
        self.Emple.ApeEmpl2.setEnabled(True)
        self.Emple.Cedu.setEnabled(True)
        self.Emple.CellNumber.setEnabled(True)
        self.Emple.Salario.setEnabled(True)
        self.Emple.FechaIngr.setEnabled(True)
        self.Emple.Cargo.setEnabled(True)

    def limpiar_datosnomi(self): #Limpiar cuadros de texto de la ventana Datos Nomina
        self.Datos.SearchName.clear()
        self.Datos.NombreEmp.clear()
        self.Datos.Cedu.clear()
        self.Datos.Cargo.clear()
        self.Datos.SalarioBase.clear()
        self.Datos.Dias.clear()
        self.Datos.Feriados.clear()
        self.Datos.Domingos.clear()
        self.Datos.Libres.clear()
        self.Datos.HExDiurnas.clear()
        self.Datos.HExNocturnas.clear()
        self.Datos.BonoNoct.clear()
        self.Datos.Reposo.clear()
        self.Datos.CT.clear()
        self.Datos.ISRL.clear()
        self.Datos.Faltan.clear()
        self.Datos.Presta.clear()
        self.Datos.Consu.clear()
        self.Datos.SearchName.setEnabled(True)
        self.Datos.Dias.setEnabled(True)
        self.Datos.Feriados.setEnabled(True)
        self.Datos.Domingos.setEnabled(True)
        self.Datos.Libres.setEnabled(True)
        self.Datos.HExDiurnas.setEnabled(True)
        self.Datos.HExNocturnas.setEnabled(True)
        self.Datos.BonoNoct.setEnabled(True)
        self.Datos.Reposo.setEnabled(True)
        self.Datos.CT.setEnabled(True)
        self.Datos.ISRL.setEnabled(True)
        self.Datos.Faltan.setEnabled(True)
        self.Datos.Presta.setEnabled(True)
        self.Datos.Consu.setEnabled(True)
        self.Datos.Periodo.setEnabled(True)
        self.Datos.Lunes.setEnabled(True)
        self.Datos.Descanso.setEnabled(True)

    def limpiar_prov(self):
        self.Prov.Rif.clear()
        self.Prov.Empresa.clear()
        self.Prov.TlfnoEmpre.clear()
        self.Prov.EmailEmpre.clear()
        self.Prov.NameVen.clear()
        self.Prov.TlfnoVende.clear()
        self.Prov.Rif.setEnabled(True)
        self.Prov.Empresa.setEnabled(True)
        self.Prov.TlfnoEmpre.setEnabled(True)
        self.Prov.EmailEmpre.setEnabled(True)
        self.Prov.NameVen.setEnabled(True)
        self.Prov.TlfnoVende.setEnabled(True)
    
    def limpiar_venta(self):
        self.Venta.SubTotal.clear()
        self.Venta.iva.clear()
        self.Venta.Pagarbs.clear()
        self.Venta.PagarD.clear()

if __name__ == '__main__': #Abrir programa
    app=QApplication(sys.argv)
    GUI = App()
    GUI.show() #Muestra el programa
    sys.exit(app.exec_()) #Permite que no se cierre el programa